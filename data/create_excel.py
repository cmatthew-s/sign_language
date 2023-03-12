# -*- coding: utf-8 -*-
"""
Created on Wed Nov  9 08:43:00 2022

@author: matthew
"""

from openpyxl import Workbook

wb = Workbook()
ws0 = wb['Sheet'] # default sheet

coord = ['x', 'y', 'z']

col_start =  1
row_start = 1
for i in range(0, 21):
    for j in coord:
        ws0.cell(row_start, col_start).value = '{}{}'.format(j, i+1)
        col_start += 1
        # print('{}{}'.format(j, i+1))

# extra cells
ws0.cell(row_start, col_start).value = 'type'
col_start += 1
ws0.cell(row_start, col_start).value = 'category'
 
# save wb
wb.save('Data_AI.xlsx')
        